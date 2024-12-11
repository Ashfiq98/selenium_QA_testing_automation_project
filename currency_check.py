# currency_check.py
import time
import os
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class CurrencySelectionBot:
    def __init__(self, url, log_callback=None):
        self.chrome_options = Options()
        self.chrome_options.add_argument("--start-maximized")
        self.chrome_options.add_argument("--disable-extensions")
        self.chrome_options.add_argument("--disable-gpu")
        self.chrome_options.add_argument("--no-sandbox")
        # self.service = Service("C:/Users/Opu/.wdm/drivers/chromedriver/win64/131.0.6778.87/chromedriver.exe")  # Update for your local driver path
        self.service = Service(ChromeDriverManager().install())
        self.driver = None
        self.url = url
        self.log_callback = log_callback
        self.results = []

    def log(self, message):
        print(message)
        if self.log_callback:
            self.log_callback(message)

    def setup_driver(self):
        self.driver = webdriver.Chrome(service=self.service, options=self.chrome_options)
        self.wait = WebDriverWait(self.driver, 30)

    def run_currency_selection_test(self):
        try:
            self.setup_driver()
            self.log("🌐 Navigating to website...")
            self.driver.get(self.url)

            self.log("🔍 Searching for currency dropdown...")
            currency_dropdown = self.wait.until(
                EC.presence_of_element_located((By.ID, 'js-currency-sort-footer'))
            )
            self.log("✅ Currency dropdown found!")

            currency_options = currency_dropdown.find_elements(
                By.XPATH, './/ul[@class="select-ul"]/li'
            )
            self.log(f"💰 Found {len(currency_options)} currency options")

            for index, option in tqdm(enumerate(currency_options, 1),
                                      total=len(currency_options),
                                      desc="Processing currencies",
                                      ncols=100, unit="option"):
                try:
                    # Extract the currency name
                    currency_raw_text = option.get_attribute('innerText')
                    # print(currency_raw_text)
                    currency_match = re.search(r'\((.*?)\)', currency_raw_text)
                    currency_text = currency_match.group(1)
                    self.log(f"\n🔄 Processing Currency Option {index}: {currency_text}")
                    
                    # Capture initial prices
                    property_prices = self.driver.find_elements(By.CLASS_NAME, 'js-price-value')
                    initial_prices = [
                        re.sub(r'^.{3}', '', price.text.strip()) for price in property_prices
                    ]
                    # initial_text = currency_text
                    self.driver.execute_script("arguments[0].click();", option)
                    time.sleep(3)  # Wait for prices to update

                    # Capture updated prices
                    updated_prices = [
                        re.sub(r'^.{3}', '', price.text.strip()) for price in property_prices
                    ]

                    # Compare initial and updated prices
                    price_changes = []
                    status = 'Pass'
                    for initial, updated in zip(initial_prices, updated_prices):
                        if initial != updated:
                            print(f"Initial : {initial:<10} --- Updated : {updated:<10}")
                            # price_changes.append(f"{initial} ➡ {updated}")
                        else:
                            status = 'Fail'  # Mark as fail if any price does not change

                    if status == 'Pass':
                        comments = f"Prices updated successfully in {currency_text}"
                        self.log(f"🟢 Property prices successfully updated in {currency_text} ")
                    else:
                        comments = "Prices did not update"
                        self.log(f"❌ Prices did not update for {currency_text}")
                    
                    self.results.append({
                        'url': self.url,
                        'currency': currency_text,
                        'status': status,
                        'comments': comments
                    })

                except Exception as e:
                    self.log(f"❌ Error with currency option {index}: {e}")
                    self.results.append({
                        'url': self.url,
                        'currency': "Unknown",
                        'status': 'Fail',
                        'comments': f"Error: {e}"
                    })
        except Exception as e:
            self.log(f"❌ Critical Test Error: {e}")
            return False
        finally:
            if self.driver:
                self.driver.quit()
        return True

    def generate_excel_report(self):
        try:
            os.makedirs('reports', exist_ok=True)
            report_file = 'reports/all_the_reports.xlsx'

            if os.path.exists(report_file):
                workbook = openpyxl.load_workbook(report_file)
                if "Currency" not in workbook.sheetnames:
                    sheet = workbook.create_sheet(title="Currency")
                else:
                    sheet = workbook["Currency"]
            else:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Currency"

            headers = ['Page URL', 'Currency', 'Status', 'Comments']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row, result in enumerate(self.results, start=2):
                sheet.cell(row=row, column=1, value=result.get('url', ''))
                sheet.cell(row=row, column=2, value=result.get('currency', ''))
                sheet.cell(row=row, column=3, value=result.get('status', ''))
                sheet.cell(row=row, column=4, value=result.get('comments', ''))

            for col in range(1, 5):
                column_letter = get_column_letter(col)
                max_length = max(len(str(cell.value or '')) for cell in sheet[get_column_letter(col)])
                sheet.column_dimensions[column_letter].width = max_length + 2

            workbook.save(report_file)
            self.log(f"✅ Excel Report generated: {report_file}")
            return report_file
        except Exception as e:
            self.log(f"❌ Error generating Excel report: {e}")
            return None


def main():
    url = "https://www.alojamiento.io/property/cabrils/BC-1178728"
    bot = CurrencySelectionBot(url)
    if bot.run_currency_selection_test():
        print("✅ Currency Selection Test Completed Successfully!")
        bot.generate_excel_report()
    else:
        print("❌ Currency Selection Test Failed")

# if __name__ == "__main__":
#     main()
