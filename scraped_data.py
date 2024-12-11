# scraped_data.py
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os


class ScrapeData:
    def __init__(self, url):
        """
        Initialize the scraper with the provided URL and set up WebDriver.
        """
        options = Options()
        options.headless = False  # Set to True to run in headless mode
        # self.service = Service("C:/Users/Opu/.wdm/drivers/chromedriver/win64/131.0.6778.87/chromedriver.exe")  # Update for your local driver path

        # self.driver = webdriver.Chrome(service=webdriver.chrome.service.Service("C:/Users/Opu/.wdm/drivers/chromedriver/win64/131.0.6778.87/chromedriver.exe"), options=options)
        self.driver = webdriver.Chrome(service=webdriver.chrome.service.Service(ChromeDriverManager().install()), options=options)
        self.url = url
        self.driver.get(self.url)

    def scrape_data(self):
        """
        Scrapes data from the ScriptData JSON object and saves it in an Excel file.
        """
        data = []

        try:
            # Execute JavaScript to get the ScriptData object
            script_data = self.driver.execute_script("return window.ScriptData;")

            if script_data:
                print("ScriptData fetched successfully!")

                # Extract required fields from the nested structure
                site_url = script_data.get("config", {}).get("SiteUrl", "")
                site_name = script_data.get("config", {}).get("SiteName", "")
                browser = script_data.get("userInfo", {}).get("Browser", "")
                country_code = script_data.get("userInfo", {}).get("CountryCode", "")
                ip = script_data.get("userInfo", {}).get("IP", "")
                campaign_id = script_data.get("pageData", {}).get("CampaignId", "")

                # Compile the data into a dictionary
                row = {
                    "SiteURL": site_url,
                    "SiteName": site_name,
                    "Browser": browser,
                    "CountryCode": country_code,
                    "IP": ip,
                    "CampaignID": campaign_id
                }
                data.append(row)

                # Save the data to an Excel file
                self.save_to_excel(data)
            else:
                print("No ScriptData found on the page.")
        except Exception as e:
            print(f"Error extracting ScriptData: {e}")

    def save_to_excel(self, data):
        """
        Saves the scraped data to an Excel file, creating or updating a specific sheet.
        """
        try:
            # Ensure the 'reports' folder exists
            os.makedirs('reports', exist_ok=True)
            report_file = 'reports/all_the_reports.xlsx'

            # Check if the file already exists
            if os.path.exists(report_file):
                # Load the existing workbook
                workbook = openpyxl.load_workbook(report_file)

                # Check if the "Script Data" sheet already exists
                if "Script Data" in workbook.sheetnames:
                    # If it exists, use that sheet
                    sheet = workbook["Script Data"]
                else:
                    # If the sheet doesn't exist, create a new one
                    sheet = workbook.create_sheet(title="Script Data")
            else:
                # If the file doesn't exist, create a new workbook and sheet
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Script Data"

            # Define headers for the sheet
            headers = ['SiteURL', 'SiteName', 'Browser', 'CountryCode', 'IP', 'CampaignID']
            # Write headers to the first row
            for col, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Write data rows
            for row_index, row_data in enumerate(data, start=2):  # Start from row 2 to leave space for headers
                sheet.cell(row=row_index, column=1, value=row_data.get('SiteURL', ''))
                sheet.cell(row=row_index, column=2, value=row_data.get('SiteName', ''))
                sheet.cell(row=row_index, column=3, value=row_data.get('Browser', ''))
                sheet.cell(row=row_index, column=4, value=row_data.get('CountryCode', ''))
                sheet.cell(row=row_index, column=5, value=row_data.get('IP', ''))
                sheet.cell(row=row_index, column=6, value=row_data.get('CampaignID', ''))

            # Adjust column widths
            for col in range(1, 7):
                column_letter = get_column_letter(col)
                max_length = 0
                for cell in sheet[column_letter]:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                sheet.column_dimensions[column_letter].width = max_length + 2

            # Save the workbook
            workbook.save(report_file)
            print(f"✅ Data saved to Excel: {report_file}")

        except Exception as e:
            print(f"❌ Error saving data to Excel: {e}")
            raise

    def close(self):
        """
        Close the WebDriver.
        """
        self.driver.quit()


# Entry point
# if __name__ == "__main__":
#     url = "https://www.alojamiento.io/"
#     scraper = ScrapeData(url)
#     scraper.scrape_data()
#     scraper.close()
