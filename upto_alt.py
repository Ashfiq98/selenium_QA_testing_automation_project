# upto_alt.py
import os
import json
import time
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


class VacationRentalTester:
    def __init__(self, url):
        """
        Initialize the tester with Chrome WebDriver
        """
        try:
            options = Options()
            options.headless = False
            # Setup Chrome WebDriver
            self.service = Service(ChromeDriverManager().install())
            # self.service = Service("C:/Users/Opu/.wdm/drivers/chromedriver/win64/131.0.6778.87/chromedriver.exe")  # For home
            self.driver = webdriver.Chrome(service=self.service)
            self.url = url
            self.results = []
            logging.info(f"Initialized tester for URL: {url}")
        except Exception as e:
            logging.error(f"Initialization error: {e}")
            raise

    def navigate(self):
        """
        Navigate to the website
        """
        try:
            self.driver.get(self.url)
            time.sleep(3)  # Wait for page load
            logging.info(f"Successfully navigated to {self.url}")
        except Exception as e:
            logging.error(f"Navigation error: {e}")
            raise

    def test_h1_tag(self):
        """
        Test H1 tag existence
        """
        try:
            h1_tags = self.driver.find_elements(By.TAG_NAME, 'h1')
            status = len(h1_tags) > 0
            self.results.append({
                'page_url': self.url,
                'testcase': 'H1 Tag Existence',
                'status': 'Pass' if status else 'Fail',
                'comments': 'H1 tag found' if status else 'No H1 tag present'
            })
        except Exception as e:
            self.results.append({
                'page_url': self.url,
                'testcase': 'H1 Tag Existence',
                'status': 'Fail',
                'comments': f'Error checking H1: {str(e)}'
            })

    def test_html_tag_sequence(self):
        """
        Test HTML heading tag sequence to ensure they follow a proper order (H1->H2->...->H6).
        """
        try:
            heading_tags = ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']
            found_tags = []
            
            # Collect heading tags with their positions
            for tag in heading_tags:
                elements = self.driver.find_elements(By.TAG_NAME, tag)
                for elem in elements:
                    found_tags.append((tag, elem.location['y']))
            
            # Sort tags by their position
            found_tags.sort(key=lambda x: x[1])  # Sort by position (y-coordinate)
            found_tag_names = [tag for tag, _ in found_tags]

            # Validate sequence
            valid_sequence = True
            for i in range(1, len(found_tag_names)):
                if heading_tags.index(found_tag_names[i - 1]) > heading_tags.index(found_tag_names[i]):
                    valid_sequence = False
                    break

            # Record results
            self.results.append({
                'page_url': self.url,
                'testcase': 'HTML Tag Sequence',
                'status': 'Pass' if valid_sequence else 'Fail',
                'comments': f'Found tags in order: {found_tag_names}' if valid_sequence else f'Invalid sequence: {found_tag_names}'
            })
        except Exception as e:
            self.results.append({
                'page_url': self.url,
                'testcase': 'HTML Tag Sequence',
                'status': 'Fail',
                'comments': f'Error checking tag sequence: {str(e)}'
            })

    def test_image_alt_attributes(self):
        """
        Test image alt attributes
        """
        try:
            images = self.driver.find_elements(By.TAG_NAME, 'img')
            failed_images = [img.get_attribute('src') for img in images if not img.get_attribute('alt')]
            
            status = len(failed_images) == 0
            self.results.append({
                'page_url': self.url,
                'testcase': 'Image Alt Attributes',
                'status': 'Pass' if status else 'Fail',
                'comments': 'All images have alt attributes' if status else f'Missing alt for: {failed_images}'
            })
        except Exception as e:
            self.results.append({
                'page_url': self.url,
                'testcase': 'Image Alt Attributes',
                'status': 'Fail',
                'comments': f'Error checking image alt attributes: {str(e)}'
            })


    def generate_excel_report(self):
        """
        Generate or overwrite Excel report of test results with formatting
        """
        try:
            # Create reports directory if it doesn't exist
            os.makedirs('reports', exist_ok=True)

            # Define the report file path
            report_file = 'reports/all_the_reports.xlsx'

            # Check if the file already exists
            if os.path.exists(report_file):
                # print("Exist...")
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                if "Test" not in workbook.sheetnames:
                    sheet = workbook.create_sheet(title="Test")
                else:
                    sheet = workbook["Test"]
            else:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet = workbook.create_sheet(title="Test")

            # Define headers
            headers = ['Page URL', 'Test Case', 'Status', 'Comments']

            # Write headers with formatting
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Write test results
            for row, result in enumerate(self.results, start=2):
                sheet.cell(row=row, column=1, value=result.get('page_url', ''))
                sheet.cell(row=row, column=2, value=result.get('testcase', ''))
                sheet.cell(row=row, column=3, value=result.get('status', ''))
                sheet.cell(row=row, column=4, value=result.get('comments', ''))

            # Auto-adjust column widths
            for col in range(1, 5):
                column_letter = get_column_letter(col)
                sheet.column_dimensions[column_letter].auto_size = True

            # Save the workbook
            workbook.save(report_file)
            logging.info(f"Excel report generated successfully: {report_file}")
            print(f"✅ Excel Report generated: {report_file}")

            return report_file

        except Exception as e:
            logging.error(f"Excel report generation error: {e}")
            print(f"❌ Error generating Excel report: {e}")
            return None

    def close(self):
        """
        Close browser
        """
        if self.driver:
            self.driver.quit()


def run_tests(url):
    tester = None
    try:
        tester = VacationRentalTester(url)
        tester.navigate()
        tester.test_h1_tag()
        tester.test_html_tag_sequence()
        tester.test_image_alt_attributes()
        # tester.test_currency_filter()  # Add currency filter test
        # tester.scrape_script_data()    # Add script data scraping test
        tester.generate_excel_report()

    except Exception as e:
        print(f"Test execution error: {e}")

    # finally:
    #     if tester:
    #         tester.close()
            
   

# Example usage
# if __name__ == "__main__":
#     url = "https://www.alojamiento.io/"  # Provide your URL
#     run_tests(url)
