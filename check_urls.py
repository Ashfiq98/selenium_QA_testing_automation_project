# check_urls.py
import os
import time
import logging
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from tqdm import tqdm


class UrlTester:
    def __init__(self, url):
        """
        Initialize the tester with Chrome WebDriver
        """
        try:
            options = Options()
            options.headless = False  # Set to True if you don't want a browser window
            self.service = Service(ChromeDriverManager().install())
            # self.service = Service("C:/Users/Opu/.wdm/drivers/chromedriver/win64/131.0.6778.87/chromedriver.exe")  # Update for your local driver path
            self.driver = webdriver.Chrome(service=self.service)
            self.url = url
            self.results = []
            logging.info(f"Initialized tester for URL: {url}")
        except Exception as e:
            logging.error(f"Initialization error: {e}")
            raise

    def navigate(self):
        """
        Navigate to the website
        """
        try:
            self.driver.get(self.url)
            time.sleep(3)  # Wait for page load
            logging.info(f"Successfully navigated to {self.url}")
        except Exception as e:
            logging.error(f"Navigation error: {e}")
            raise

    def check_all_urls(self):
        """
        Check the status codes of all the links on the webpage and show progress.
        """
        try:
            self.driver.get(self.url)
            time.sleep(3)  # Wait for page load
            links = self.driver.find_elements(By.TAG_NAME, 'a')

            # Log total number of links found
            logging.info(f"Found {len(links)} links on the page.")
            print(f"✅ Found {len(links)} links on the page.")

            # Use tqdm for progress bar
            for link in tqdm(links, desc="Checking URLs", unit="URL"):
                href = link.get_attribute('href')
                if href:  # Only check links that have an href attribute
                    print(f"⏳ Checking URL: {href}")
                    logging.info(f"Checking URL: {href}")
                    # response = requests.get(href)
                    try:
                        response = requests.get(href)
                        if response.status_code == 404:
                            print(f"❌ 404 Not Found: {href}")
                            self.results.append({
                                'page_url': href,
                                'testcase': 'URL Status Code',
                                'status': 'Fail',
                                'comments': '404 Not Found'
                            })
                        else:
                            print(f"✅ URL Status Code: {response.status_code}")
                            # self.results.append({
                            #     'page_url': href,
                            #     'testcase': 'URL Status Code',
                            #     'status': 'Pass',
                            #     'comments': f'Status Code: {response.status_code}'
                            # })
                    except Exception as e:
                        print(f"❌ Error checking URL: {href}")
                        self.results.append({
                            'page_url': href,
                            'testcase': 'URL Status Code',
                            'status': 'Fail',
                            'comments': f'Error: {str(e)}'
                        })
                    time.sleep(1)  # Optional delay for better visibility

            # Generate Excel report after checking all URLs
            # self.generate_excel_report()

        except Exception as e:
            logging.error(f"Error checking URLs on the page: {e}")
            raise

    def generate_excel_report(self):
        """
        Generate an Excel report for test results.
        """
        try:
            os.makedirs('reports', exist_ok=True)
            report_file = 'reports/all_the_reports.xlsx'
            if os.path.exists(report_file):
                print("Exist...")
                # workbook = openpyxl.Workbook()
                # sheet = workbook.active
                workbook = openpyxl.load_workbook(report_file)
                if "Test" not in workbook.sheetnames:
                    sheet = workbook.create_sheet(title="Test")
                else:
                    sheet = workbook["Test"]
            else:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "Test"
            # Define headers
            if sheet.max_row == 1:
                headers = ['Page URL', 'Test Case', 'Status', 'Comments']
                for col, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=1, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Write results
            for row, result in enumerate(self.results, start=5):
                sheet.cell(row=row, column=1, value=result['page_url'])
                sheet.cell(row=row, column=2, value=result['testcase'])
                sheet.cell(row=row, column=3, value=result['status'])
                sheet.cell(row=row, column=4, value=result['comments'])

            # Adjust column widths
            for col in range(1, 5):
                column_letter = get_column_letter(col)
                sheet.column_dimensions[column_letter].width = 30

            # Save the workbook
            workbook.save(report_file)
            print(f"✅ URL Test Results saved to: {report_file}")
            return report_file

        except Exception as e:
            print(f"❌ Error generating Excel report: {e}")
            raise

    def close(self):
        """
        Close the WebDriver.
        """
        if self.driver:
            self.driver.quit()


def run_tests_url(url):
    """
    Run all the tests for the given URL.
    """
    tester = None
    try:
        tester = UrlTester(url)
        tester.navigate()
        tester.check_all_urls()  # Check all URLs and generate the report
        tester.generate_excel_report()
    except Exception as e:
        print(f"Test execution error: {e}")
    # finally:
    #     if tester:
    #         tester.close()


# Entry point for running tests
# if __name__ == "__main__":
#     test_url = "https://www.alojamiento.io/"
#     run_tests_url(test_url)
